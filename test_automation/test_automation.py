from playwright.sync_api import sync_playwright
import time
import os
import argparse
import re
from pathlib import Path
import sys
import openpyxl
from openpyxl.cell.cell import MergedCell

ROOT_DIR = Path(__file__).resolve().parent.parent
TESTS_DIR = ROOT_DIR / "test_automation"
DEFAULT_EXCEL = str(TESTS_DIR / "Assignment 1 - Test cases.xlsx")
DEFAULT_SHEET_NAME = " Test cases"
DEFAULT_FRONTEND_URL = os.getenv("FRONTEND_URL", "https://www.pixelssuite.com/chat-translator")

def _configure_stdout():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass

def _resolve_path(p):
    if not p:
        return None
    path = Path(p)
    if path.is_absolute():
        return str(path)
    root_candidate = (ROOT_DIR / path).resolve()
    if root_candidate.exists():
        return str(root_candidate)
    tests_candidate = (TESTS_DIR / path).resolve()
    if tests_candidate.exists():
        return str(tests_candidate)
    return str(root_candidate)

def _norm(v):
    return re.sub(r"[^a-z0-9]+", "", str(v or "").strip().lower())

def _find_header_row(ws):
    for r in range(1, min(ws.max_row, 30) + 1):
        vals = [_norm(ws.cell(r,c).value) for c in range(1, min(ws.max_column, 12) + 1)]
        if "input" in vals and ("actualoutput" in vals or "expectedoutput" in vals):
            return r
    return 1

def _find_exact_col(ws, header_row, header_name, fallback_col):
    target = _norm(header_name)
    for c in range(1, ws.max_column + 1):
        if _norm(ws.cell(header_row,c).value) == target:
            return c
    return fallback_col

def _merged_top_left_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    return cell

def _is_top_left_of_merged_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return True
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row == row and rng.min_col == col
    return True

def _set_cell(ws, row, col, value):
    _merged_top_left_cell(ws, row, col).value = value

def _dismiss_overlays(page):
    for text in ["Accept", "I Agree", "Agree", "OK", "Got it", "Accept all"]:
        try:
            btn = page.get_by_role("button", name=re.compile(rf"^{re.escape(text)}$", re.I)).first
            if btn.is_visible():
                btn.click(timeout=1500)
                page.wait_for_timeout(300)
        except Exception:
            pass

def _clear_textarea(page, locator):
    try:
        locator.click(timeout=2000)
        page.keyboard.press("Control+A")
        page.keyboard.press("Backspace")
    except Exception:
        pass
    try:
        locator.fill("")
    except Exception:
        pass
    try:
        locator.evaluate("""el => { el.value=''; el.dispatchEvent(new Event('input', {bubbles:true})); }""")
    except Exception:
        pass

def _type_input(page, input_locator, text, delay):
    _clear_textarea(page, input_locator)
    if delay and delay > 0:
        input_locator.click(timeout=2000)
        input_locator.type(text, delay=delay)
    else:
        input_locator.fill(text)

def _read_output(output_locator):
    for getter in [lambda: output_locator.input_value(), lambda: output_locator.inner_text(), lambda: output_locator.text_content(), lambda: output_locator.evaluate("el => ('value' in el ? el.value : el.textContent)")]:
        try:
            v = getter()
            if v is not None and str(v).strip():
                return str(v).strip()
        except Exception:
            pass
    return ""

def _find_chat_locators(page, timeout_ms):
    deadline = time.time() + timeout_ms/1000
    while time.time() < deadline:
        _dismiss_overlays(page)
        try:
            input_box = page.locator('textarea[placeholder*="English"]').first
            output_box = page.locator('textarea[placeholder*="Sinhala"]').first
            if input_box.count() and output_box.count() and input_box.is_visible() and output_box.is_visible():
                btn = page.get_by_role("button", name=re.compile(r"^Transliterate$", re.I)).first
                return input_box, output_box, btn
        except Exception:
            pass
        try:
            visible = []
            for i in range(page.locator("textarea").count()):
                loc = page.locator("textarea").nth(i)
                if loc.is_visible():
                    visible.append(loc)
            if len(visible) >= 2:
                btn = page.get_by_role("button", name=re.compile(r"^Transliterate$", re.I)).first
                return visible[0], visible[1], btn
        except Exception:
            pass
        page.wait_for_timeout(500)
    raise RuntimeError("Could not find input/output textarea boxes.")

def _parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--excel", default=DEFAULT_EXCEL)
    p.add_argument("--sheet", default=DEFAULT_SHEET_NAME)
    p.add_argument("--url", default=DEFAULT_FRONTEND_URL)
    p.add_argument("--output", default=None)
    p.add_argument("--wait-ms", type=int, default=5000)
    p.add_argument("--retries", type=int, default=8)
    p.add_argument("--retry-wait-ms", type=int, default=1000)
    p.add_argument("--type-delay-ms", type=int, default=80)
    p.add_argument("--timeout-ms", type=int, default=60000)
    p.add_argument("--slow-mo-ms", type=int, default=200)
    p.add_argument("--save-every", type=int, default=1)
    p.add_argument("--headless", action="store_true")
    p.add_argument("--keep-open", action="store_true")
    return p.parse_args()

def run_test():
    _configure_stdout()
    args = _parse_args()
    args.excel = _resolve_path(args.excel)
    args.output = _resolve_path(args.output) if args.output else args.excel
    if not args.excel or not os.path.exists(args.excel):
        print(f"Error: Excel file not found: {args.excel}")
        return
    wb = openpyxl.load_workbook(args.excel)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb.active
    header_row = _find_header_row(ws)

    # IMPORTANT FIX: ONLY column named exactly Input is tested. Category/type/evidence columns are ignored.
    input_col = _find_exact_col(ws, header_row, "Input", 3)
    expected_col = _find_exact_col(ws, header_row, "Expected output", 4)
    actual_col = _find_exact_col(ws, header_row, "Actual output", 5)
    status_col = _find_exact_col(ws, header_row, "Status", 6)

    print(f"Using Excel: {args.excel}")
    print(f"Header row: {header_row}; Input column: {input_col}; Actual output column: {actual_col}")
    print("Only values from the Input column will be typed into the website.")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=args.headless, slow_mo=max(0, args.slow_mo_ms))
        page = browser.new_page()
        page.set_default_timeout(max(1000, args.timeout_ms))
        page.goto(args.url, wait_until="domcontentloaded")
        try:
            page.wait_for_load_state("networkidle", timeout=args.timeout_ms)
        except Exception:
            pass
        input_box, output_box, btn = _find_chat_locators(page, args.timeout_ms)
        processed = 0
        for r in range(header_row + 1, ws.max_row + 1):
            if not _is_top_left_of_merged_cell(ws, r, input_col):
                continue
            input_value = _merged_top_left_cell(ws, r, input_col).value
            singlish_input = str(input_value).strip() if input_value is not None else ""
            if not singlish_input:
                continue
            print(f"Testing row {r}: {singlish_input}")
            try:
                _dismiss_overlays(page)
                previous = _read_output(output_box)
                _type_input(page, input_box, singlish_input, args.type_delay_ms)
                try:
                    btn.click()
                except Exception:
                    pass
                page.wait_for_timeout(max(0, args.wait_ms))
                actual = ""
                for _ in range(max(1, args.retries)):
                    current = _read_output(output_box)
                    if current and current != previous:
                        actual = current
                        break
                    page.wait_for_timeout(max(0, args.retry_wait_ms))
                if not actual:
                    actual = _read_output(output_box)
                _set_cell(ws, r, actual_col, actual)
                expected = _merged_top_left_cell(ws, r, expected_col).value
                expected = str(expected).strip() if expected is not None else ""
                _set_cell(ws, r, status_col, "FAIL" if expected and actual != expected else ("PASS" if expected else "COLLECTED"))
                processed += 1
                print(f"  -> Actual output saved. Status: {_merged_top_left_cell(ws, r, status_col).value}")
                if args.save_every and processed % args.save_every == 0:
                    wb.save(args.output)
            except Exception as e:
                _set_cell(ws, r, status_col, "UI Error")
                print(f"  -> UI Error: {e}")
                if args.save_every:
                    wb.save(args.output)
        wb.save(args.output)
        if args.keep_open and not args.headless:
            print("Keeping browser open. Press CTRL+C to stop.")
            try:
                while True:
                    page.wait_for_timeout(1000)
            except KeyboardInterrupt:
                wb.save(args.output)
        browser.close()
    print(f"Done. Results saved to {args.output}")

if __name__ == "__main__":
    run_test()
